import serial
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import os

ser = serial.Serial('/dev/ttyS0', 115200, timeout=1)

base_directory = "/home/pc/Data/"

if not os.path.exists(base_directory):
    os.makedirs(base_directory)

current_day = datetime.now().strftime("%Y-%m-%d")
current_directory = os.path.join(base_directory, current_day)
if not os.path.exists(current_directory):
    os.makedirs(current_directory)

start_time = datetime.now()
filename = os.path.join(current_directory, f"data_log_{start_time.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")

# Function to create a new workbook
def create_new_workbook(file_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data Log"
    sheet.append(["Timestamp", "Temperature", "Humidity", "Air Quality", "Water Temp.", "Water Lvl.",
                  "Hour", "Min", "Sec", "FlagIR", "FlagVent", "FlagLight", "FlagHeat"])
    workbook.save(file_path)

# Create initial file
create_new_workbook(filename)

while True:
    now = datetime.now()

    # Check if a new day has started
    if now.strftime("%Y-%m-%d") != current_day:
        current_day = now.strftime("%Y-%m-%d")
        current_directory = os.path.join(base_directory, current_day)
        if not os.path.exists(current_directory):
            os.makedirs(current_directory)

    # Check if 6 hours have passed since the last file creation
    if (now - start_time) >= timedelta(hours=6):
        start_time = now
        filename = os.path.join(current_directory, f"data_log_{start_time.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")
        create_new_workbook(filename)

    # Read data from serial port
    if ser.in_waiting > 0:
        line = ser.readline().decode('utf-8').rstrip()
        print("Received:", line)
        data = line.split(",")
        timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
        data.insert(0, timestamp)

        try:
            workbook = load_workbook(filename)
            sheet = workbook.active
            sheet.append(data)
            workbook.save(filename)
        except Exception as e:
            print("Failed to save data to Excel file:", e)
