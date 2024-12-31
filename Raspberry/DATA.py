import serial
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

ser = serial.Serial('/dev/ttyS0', 115200, timeout=1)

save_directory = "/home/pc/Data/"

if not os.path.exists(save_directory):
    os.makedirs(save_directory)

current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
filename = os.path.join(save_directory, f"data_log_{current_time}.xlsx")

try:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data Log"
    sheet.append(["Timestamp", "Temperature", "Humidity", "Air Quality", "Water Temp.", "Water Lvl.", 
                  "Hour", "Min", "Sec", "FlagIR", "FlagVent", "FlagLight", "FlagHeat"])
    workbook.save(filename)
except Exception as e:
    print("Excel dosyas? olu?turulamad?:", e)

while True:
    if ser.in_waiting > 0:
        line = ser.readline().decode('utf-8').rstrip()
        print("Received:", line)
        data = line.split(",")
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        data.insert(0, timestamp)
        try:
            workbook = load_workbook(filename)
            sheet = workbook.active
            sheet.append(data)
            workbook.save(filename)
        except Exception as e:
            print("Veriler Excel dosyas?na kaydedilemedi:", e)
